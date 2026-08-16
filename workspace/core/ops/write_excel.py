"""
MEKA Core SDK

Domain:
    Filesystem

Component:
    Operations - write_excel()

Purpose:
    Create or update one sheet of an Excel workbook from a 2D array of
    values, preserving any other sheets already in the file.
"""

"""
MEKA Metadata

Domain:
    filesystem

Component:
    ops.write_excel

Public API:
    write_excel

Dependencies:
    openpyxl
    pathlib
    workspace.internal.path

Thread Safe:
    yes

Pure:
    no
"""

# ==========================================================================
# Imports
# ==========================================================================

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from workspace.internal.path import resolve as resolve_path


# ==========================================================================
# Public API
# ==========================================================================

def write_excel(
    root: Path,
    path: str | Path,
    data: list[list[object]],
    *,
    sheet: str = "Sheet1",
    start_cell: str = "A1",
) -> None:
    """Write a 2D array of values into one sheet, starting at ``start_cell``.

    Parameters
    ----------
    root : Path
        Workspace root.

    path : str | Path
        Relative path inside the workspace for the workbook.

    data : list[list[object]]
        Rows of cell values to write.

    sheet : str, optional
        Sheet to write into. Created if it does not exist. Defaults to
        "Sheet1".

    start_cell : str, optional
        Top-left cell where writing begins, e.g. "A1".

    Notes
    -----
    If the workbook already exists, only ``sheet`` is replaced or created —
    every other sheet in the workbook is preserved untouched.
    """
    target = resolve_path(root, path)

    if target.exists():
        workbook = load_workbook(target)
        if sheet in workbook.sheetnames:
            del workbook[sheet]
        worksheet = workbook.create_sheet(sheet)
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet

    start_column_letter, start_row = coordinate_from_string(start_cell)
    start_col = column_index_from_string(start_column_letter)

    for row_offset, row in enumerate(data):
        for col_offset, value in enumerate(row):
            worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)

    workbook.save(target)
    workbook.close()
