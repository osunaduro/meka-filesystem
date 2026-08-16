"""
MEKA Core SDK

Domain:
    Filesystem

Component:
    Operations - read_excel()

Purpose:
    Read a rectangular range of cell values from one sheet of an Excel
    workbook.
"""

"""
MEKA Metadata

Domain:
    filesystem

Component:
    ops.read_excel

Public API:
    read_excel

Dependencies:
    datetime
    openpyxl
    pathlib
    workspace.core.errors
    workspace.core.models
    workspace.internal.path

Thread Safe:
    yes

Pure:
    no
"""

# ==========================================================================
# Imports
# ==========================================================================

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

from workspace.core.errors import SheetNotFoundError
from workspace.core.models import ExcelSheetResult
from workspace.internal.path import resolve as resolve_path


# ==========================================================================
# Internal helpers
# ==========================================================================

def _json_safe(value: object) -> object:
    """Convert a cell value into something that serializes cleanly as JSON."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


# ==========================================================================
# Public API
# ==========================================================================

def read_excel(
    root: Path,
    path: str | Path,
    *,
    sheet: str | None = None,
    cell_range: str | None = None,
) -> ExcelSheetResult:
    """Read a rectangular range of cell values from one sheet.

    Parameters
    ----------
    root : Path
        Workspace root.

    path : str | Path
        Relative path inside the workspace to an .xlsx workbook.

    sheet : str | None, optional
        Sheet name to read. Defaults to the workbook's active sheet.

    cell_range : str | None, optional
        A range like "A1:D100", relative to the selected sheet. Omit to
        read the sheet's entire used range.

    Returns
    -------
    ExcelSheetResult
        ``data`` is a 2D array (rows of cell values); ``sheet_names`` lists
        every sheet in the workbook.

    Raises
    ------
    SheetNotFoundError
        ``sheet`` was given but does not exist in the workbook.
    """
    target = resolve_path(root, path)
    workbook = load_workbook(target, data_only=True, read_only=True)

    try:
        sheet_names = list(workbook.sheetnames)

        if sheet is None:
            worksheet = workbook[workbook.active.title]
        elif sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
        else:
            raise SheetNotFoundError(f"Sheet '{sheet}' does not exist in {path}.")

        if cell_range:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        else:
            min_col, min_row = 1, 1
            max_col, max_row = worksheet.max_column, worksheet.max_row

        data: list[list[object]] = [
            [_json_safe(cell) for cell in row]
            for row in worksheet.iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
            )
        ]

        return ExcelSheetResult(
            path=target,
            sheet_name=worksheet.title,
            sheet_names=sheet_names,
            data=data,
            total_rows=max(0, max_row - min_row + 1),
            total_cols=max(0, max_col - min_col + 1),
        )
    finally:
        workbook.close()
